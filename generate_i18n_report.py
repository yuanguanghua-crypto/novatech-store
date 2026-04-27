"""生成 LabProGlobal i18n 检测报告 Excel"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ========== 读取 JSON 报告 ==========
with open(r"E:\novatech-store\i18n_report.json", encoding="utf-8") as f:
    report = json.load(f)

summary = report["summary"]
# 修正 locale_count（JSON 中正则提取有误）
summary["supported_locales"] = 7  # en/zh/es/ja/hi/ar/pt
pages = report["pages_needing_review"]
components = report["components_needing_review"]

# ========== 样式定义 ==========
def header_style(cell):
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor="1F4E79")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def ok_style(cell):
    cell.font = Font(name="Arial", color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", fgColor="375623")
    cell.alignment = Alignment(horizontal="center", vertical="center")

def warning_style(cell):
    cell.font = Font(name="Arial", color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", fgColor="C65911")
    cell.alignment = Alignment(horizontal="center", vertical="center")

def error_style(cell):
    cell.font = Font(name="Arial", color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", fgColor="C00000")
    cell.alignment = Alignment(horizontal="center", vertical="center")

def delegated_style(cell):
    cell.font = Font(name="Arial", size=10)
    cell.fill = PatternFill("solid", fgColor="D6E4F0")
    cell.alignment = Alignment(horizontal="center", vertical="center")

def title_style(cell, row):
    cell.font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.fill = PatternFill("solid", fgColor="DEEAF1")

def subtitle_style(cell):
    cell.font = Font(name="Arial", bold=True, size=11, color="1F4E79")
    cell.fill = PatternFill("solid", fgColor="EBF3FB")

def thin_border():
    thin = Side(style="thin", color="B8CCE4")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def set_row_border(ws, row, cols):
    for col in range(1, cols + 1):
        ws.cell(row=row, column=col).border = thin_border()

# ========== 创建工作簿 ==========
wb = Workbook()
wb.remove(wb.active)

# =============================================
# Sheet 1: 概览摘要
# =============================================
ws1 = wb.create_sheet("概览摘要")
ws1.column_dimensions["A"].width = 30
ws1.column_dimensions["B"].width = 20
ws1.column_dimensions["C"].width = 30
ws1.column_dimensions["D"].width = 20

# 标题
ws1.merge_cells("A1:D1")
ws1["A1"] = "LabProGlobal i18n 多语言检测报告"
title_style(ws1["A1"], 1)
ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 36

ws1.merge_cells("A2:D2")
ws1["A2"] = f"翻译键总数: {summary['translation_keys_count']} | 支持语言: {summary['supported_locales']} 种 (EN/ZH/ES/JA/HI/AR/PT)"
ws1["A2"].font = Font(name="Arial", italic=True, size=10, color="595959")
ws1["A2"].alignment = Alignment(horizontal="center")

# 统计卡片
ws1.row_dimensions[4].height = 22
headers = ["类别", "总计", "已多语言", "待处理"]
for col, h in enumerate(headers, 1):
    c = ws1.cell(row=4, column=col, value=h)
    header_style(c)

ws1.row_dimensions[5].height = 20
ws1["A5"] = "页面 (app/)"
ws1["B5"] = summary["total_pages"]
ws1["C5"] = summary["pages_with_i18n"]
ws1["D5"] = summary["pages_missing_i18n"]
for col in range(1, 5):
    ws1.cell(row=5, column=col).font = Font(name="Arial", size=10)
    ws1.cell(row=5, column=col).border = thin_border()
    ws1.cell(row=5, column=col).alignment = Alignment(horizontal="center", vertical="center")

ws1.row_dimensions[6].height = 20
ws1["A6"] = "组件 (components/)"
ws1["B6"] = summary["total_components"]
ws1["C6"] = summary["components_with_i18n"]
ws1["D6"] = summary["components_missing_i18n"]
for col in range(1, 5):
    ws1.cell(row=6, column=col).font = Font(name="Arial", size=10)
    ws1.cell(row=6, column=col).border = thin_border()
    ws1.cell(row=6, column=col).alignment = Alignment(horizontal="center", vertical="center")

# 状态说明
ws1.merge_cells("A8:D8")
ws1["A8"] = "状态说明"
subtitle_style(ws1["A8"])

legend = [
    ("[OK] OK", "已实现多语言"),
    ("[partial] 部分硬编码", "已使用 i18n 但仍有少量硬编码英文"),
    ("[->] 委托 Client", "委托给 Client Component，需检查该组件"),
    ("[X] 缺失 i18n", "完全缺失多语言支持"),
]
for i, (status, desc) in enumerate(legend, 9):
    ws1.cell(row=i, column=1, value=status).font = Font(name="Arial", bold=True, size=10)
    ws1.cell(row=i, column=2, value=desc).font = Font(name="Arial", size=10)
    ws1.cell(row=i, column=1).border = thin_border()
    ws1.cell(row=i, column=2).border = thin_border()
    ws1.row_dimensions[i].height = 18

# =============================================
# Sheet 2: 页面 i18n 状态
# =============================================
ws2 = wb.create_sheet("页面状态")
ws2.column_dimensions["A"].width = 45
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 12
ws2.column_dimensions["D"].width = 14
ws2.column_dimensions["E"].width = 50

ws2.merge_cells("A1:E1")
ws2["A1"] = f"页面 i18n 状态总览 ({summary['total_pages']} 个页面)"
title_style(ws2["A1"], 1)
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 32

headers2 = ["页面路径", "客户端?", "有 i18n?", "状态", "主要问题"]
for col, h in enumerate(headers2, 1):
    c = ws2.cell(row=2, column=col, value=h)
    header_style(c)
ws2.row_dimensions[2].height = 22

# 状态映射颜色
status_map = {
    "[OK] OK": ok_style,
    "[!] 部分硬编码": warning_style,
    "[->] 委托给 Client": delegated_style,
    "[X] 缺失 i18n": error_style,
}

for i, p in enumerate(pages, 3):
    issues_text = "; ".join([f"L{r['lineno']} {r['type']}: {r['text']}" for r in p["issues"]]) if p["issues"] else "无"

    ws2.cell(row=i, column=1, value=p["path"]).font = Font(name="Arial", size=10)
    ws2.cell(row=i, column=2, value="是" if p["is_client"] else "否").font = Font(name="Arial", size=10)
    ws2.cell(row=i, column=3, value="是" if p["has_i18n"] else "否").font = Font(name="Arial", size=10)
    ws2.cell(row=i, column=4, value=p["status"]).font = Font(name="Arial", size=10)
    ws2.cell(row=i, column=5, value=issues_text).font = Font(name="Arial", size=9, color="595959")

    for col in range(1, 6):
        ws2.cell(row=i, column=col).border = thin_border()
        ws2.cell(row=i, column=col).alignment = Alignment(vertical="center", wrap_text=True)

    # 状态列着色
    status_map.get(p["status"], delegated_style)(ws2.cell(row=i, column=4))
    ws2.row_dimensions[i].height = 30

# =============================================
# Sheet 3: 组件 i18n 状态
# =============================================
ws3 = wb.create_sheet("组件状态")
ws3.column_dimensions["A"].width = 45
ws3.column_dimensions["B"].width = 12
ws3.column_dimensions["C"].width = 12
ws3.column_dimensions["D"].width = 14
ws3.column_dimensions["E"].width = 50

ws3.merge_cells("A1:E1")
ws3["A1"] = f"组件 i18n 状态总览 ({summary['total_components']} 个组件)"
title_style(ws3["A1"], 1)
ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 32

for col, h in enumerate(headers2, 1):
    c = ws3.cell(row=2, column=col, value=h)
    header_style(c)
ws3.row_dimensions[2].height = 22

for i, c in enumerate(components, 3):
    issues_text = "; ".join([f"L{r['lineno']} {r['type']}: {r['text']}" for r in c["issues"]]) if c["issues"] else "无"

    ws3.cell(row=i, column=1, value=c["path"]).font = Font(name="Arial", size=10)
    ws3.cell(row=i, column=2, value="是" if c["is_client"] else "否").font = Font(name="Arial", size=10)
    ws3.cell(row=i, column=3, value="是" if c["has_i18n"] else "否").font = Font(name="Arial", size=10)
    ws3.cell(row=i, column=4, value=c["status"]).font = Font(name="Arial", size=10)
    ws3.cell(row=i, column=5, value=issues_text).font = Font(name="Arial", size=9, color="595959")

    for col in range(1, 6):
        ws3.cell(row=i, column=col).border = thin_border()
        ws3.cell(row=i, column=col).alignment = Alignment(vertical="center", wrap_text=True)

    status_map.get(c["status"], delegated_style)(ws3.cell(row=i, column=4))
    ws3.row_dimensions[i].height = 30

# =============================================
# Sheet 4: 待修复清单（优先级排序）
# =============================================
ws4 = wb.create_sheet("待修复清单")
ws4.column_dimensions["A"].width = 10
ws4.column_dimensions["B"].width = 20
ws4.column_dimensions["C"].width = 45
ws4.column_dimensions["D"].width = 18
ws4.column_dimensions["E"].width = 40

ws4.merge_cells("A1:E1")
ws4["A1"] = "i18n 待修复清单（按优先级排列）"
title_style(ws4["A1"], 1)
ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 32

headers4 = ["优先级", "类型", "文件路径", "问题数量", "主要问题"]
for col, h in enumerate(headers4, 1):
    c = ws4.cell(row=2, column=col, value=h)
    header_style(c)
ws4.row_dimensions[2].height = 22

# 优先级排序：先页面后组件，按缺失程度排序
priority_items = []
for p in pages:
    if p["status"] in ("[X] 缺失 i18n",):
        priority_items.append(("1-紧急", "页面", p["path"], p["hardcoded_count"], p["issues"]))
for c in components:
    if c["status"] in ("[X] Client无i18n",):
        priority_items.append(("2-紧急", "组件", c["path"], c["hardcoded_count"], c["issues"]))
for p in pages:
    if p["status"] == "[!] 部分硬编码":
        priority_items.append(("3-次要", "页面", p["path"], p["hardcoded_count"], p["issues"]))
for c in components:
    if c["status"] == "[!] 部分硬编码":
        priority_items.append(("3-次要", "组件", c["path"], c["hardcoded_count"], c["issues"]))

# 优先级着色
priority_colors = {
    "1-紧急": "FFD7D7",  # 浅红
    "2-紧急": "FFE0CC",  # 浅橙
    "3-次要": "FFFACD",  # 浅黄
}

for i, (priority, ptype, path, count, issues) in enumerate(priority_items, 3):
    issues_text = "; ".join([f"L{r['lineno']} {r['text']}" for r in issues]) if issues else "无"
    ws4.cell(row=i, column=1, value=priority)
    ws4.cell(row=i, column=2, value=ptype)
    ws4.cell(row=i, column=3, value=path)
    ws4.cell(row=i, column=4, value=count)
    ws4.cell(row=i, column=5, value=issues_text)

    for col in range(1, 6):
        ws4.cell(row=i, column=col).border = thin_border()
        ws4.cell(row=i, column=col).alignment = Alignment(vertical="center", wrap_text=True)
        ws4.cell(row=i, column=col).font = Font(name="Arial", size=10)

    fill_color = priority_colors.get(priority, "FFFFFF")
    for col in range(1, 6):
        ws4.cell(row=i, column=col).fill = PatternFill("solid", fgColor=fill_color)
    ws4.row_dimensions[i].height = 30

# 保存
output_path = r"E:\novatech-store\i18n_audit_report.xlsx"
wb.save(output_path)
print(f"Excel 报告已生成: {output_path}")
